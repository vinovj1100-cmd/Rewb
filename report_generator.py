"""
Report Generator v4.4 — PDF, Excel, CSV generation
"""
import csv, io, json
from typing import Dict, List, Any, Optional
from datetime import datetime
from dataclasses import dataclass

@dataclass
class ReportConfig:
    title: str
    format: str  # pdf, excel, csv, json
    columns: List[str]
    data: List[Dict[str, Any]]
    summary: Optional[Dict] = None

class ReportGenerator:
    def __init__(self):
        self.templates = {
            "inventory": ["sku", "name", "location", "quantity", "status"],
            "orders": ["order_id", "customer", "status", "items", "value", "created"],
            "operations": ["operation", "zone", "operator", "qty", "time", "efficiency"],
            "audit": ["timestamp", "user", "action", "entity", "details"]
        }

    def generate_csv(self, config: ReportConfig) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=config.columns)
        writer.writeheader()
        for row in config.data:
            writer.writerow({k: row.get(k, "") for k in config.columns})
        return output.getvalue()

    def generate_json(self, config: ReportConfig) -> str:
        payload = {
            "title": config.title,
            "generated_at": datetime.now().isoformat(),
            "columns": config.columns,
            "data": config.data,
            "summary": config.summary or {}
        }
        return json.dumps(payload, indent=2, default=str)

    def generate_excel(self, config: ReportConfig) -> bytes:
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Title
            ws.merge_cells("A1:" + chr(64 + len(config.columns)) + "1")
            ws["A1"] = config.title
            ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Headers
            for col_idx, col_name in enumerate(config.columns, 1):
                cell = ws.cell(row=3, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row_idx, row in enumerate(config.data, 4):
                for col_idx, col_name in enumerate(config.columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ImportError:
            # Fallback to CSV
            return self.generate_csv(config).encode("utf-8")

    def generate_pdf(self, config: ReportConfig) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()

            elements.append(Paragraph(f"<b>{config.title}</b>", styles["Title"]))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
            elements.append(Spacer(1, 12))

            table_data = [config.columns]
            for row in config.data:
                table_data.append([str(row.get(c, "")) for c in config.columns])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F9FA")),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(table)
            doc.build(elements)
            return buf.getvalue()
        except ImportError:
            return b"PDF generation requires reportlab. Install with: pip install reportlab"

    def generate(self, config: ReportConfig) -> tuple:
        if config.format.lower() == "csv":
            return self.generate_csv(config), "text/csv", f"{config.title}.csv"
        elif config.format.lower() == "json":
            return self.generate_json(config), "application/json", f"{config.title}.json"
        elif config.format.lower() == "excel":
            return self.generate_excel(config), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{config.title}.xlsx"
        elif config.format.lower() == "pdf":
            return self.generate_pdf(config), "application/pdf", f"{config.title}.pdf"
        else:
            return self.generate_csv(config), "text/csv", f"{config.title}.csv"
